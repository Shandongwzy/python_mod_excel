import os
import sys
import re
import shutil
import logging
import string
import pandas as pd
import xlrd
from openpyxl import load_workbook
from xlutils.copy import copy as xl_copy

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def get_base_path():
    """Directory of exe when frozen, otherwise directory of script."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def excel_col_to_index(col):
    """Convert Excel column letters (e.g. 'A', 'H', 'AA') to 0-based index."""
    if col is None:
        raise ValueError("Empty column spec")
    col = str(col).strip().upper()
    if not col:
        raise ValueError("Empty column spec")
    num = 0
    for ch in col:
        if ch < 'A' or ch > 'Z':
            raise ValueError(f"Invalid column letter: {col}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num - 1

def normalize_sheet_spec(v):
    """
    Convert sheet spec from rules to either int (0-based index) or str (sheet name).
    Accept numeric floats (0.0) or numeric strings -> int.
    """
    if pd.isna(v):
        return 0
    # if it's float or int-like
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, (int,)):
        return int(v)
    s = str(v).strip()
    if s == "":
        return 0
    # numeric string?
    if s.isdigit():
        return int(s)
    return s  # treat as name

def read_rules(rules_path):
    """
    Read rules file into a pandas DataFrame.
    - If .xls: read with xlrd directly (avoid pandas/xlrd version checks).
    - If .xlsx: read with pandas+openpyxl.
    """
    ext = os.path.splitext(rules_path)[1].lower()
    if ext == ".xls":
        # Read using xlrd and convert to DataFrame to avoid pandas' xlrd requirements.
        book = xlrd.open_workbook(rules_path)
        sheet = book.sheet_by_index(0)
        rows = [sheet.row_values(r) for r in range(sheet.nrows)]
        if not rows:
            return pd.DataFrame()
        header = [str(h) if h is not None else "" for h in rows[0]]
        data = []
        for r in rows[1:]:
            # normalize row length to header length
            if len(r) < len(header):
                r = list(r) + [None] * (len(header) - len(r))
            elif len(r) > len(header):
                r = r[:len(header)]
            data.append(r)
        df = pd.DataFrame(data, columns=header)
        return df
    elif ext == ".xlsx":
        return pd.read_excel(rules_path, header=0, engine="openpyxl")
    else:
        raise ValueError(f"Unsupported rules file extension: {ext}")

def find_header_index_xls(sheet, header_name):
    """Return 0-based column index by header string in xlrd sheet header row (row 0)."""
    headers = [str(h).strip() if h is not None else "" for h in sheet.row_values(0)]
    try:
        return headers.index(str(header_name))
    except ValueError:
        return None

def process_xls(input_path, sheet_spec, regex_col_spec, changes):
    """
    Use xlrd + xlutils.copy to modify an .xls workbook, preserving formatting where possible.
    sheet_spec: int index or string name
    regex_col_spec: either header name, Excel letter 'H', or numeric index
    changes: dict with 'regex' (pattern) and 'changes' dict mapping column-spec -> value
    """
    book = xlrd.open_workbook(input_path, formatting_info=True)
    # resolve sheet index
    if isinstance(sheet_spec, int):
        sheet_idx = sheet_spec
    else:
        # sheet_spec is name
        try:
            sheet_idx = book.sheet_names().index(str(sheet_spec))
        except ValueError:
            raise ValueError(f"Sheet name '{sheet_spec}' not found in {input_path}")
    sheet = book.sheet_by_index(sheet_idx)
    wb_copy = xl_copy(book)
    sheet_writable = wb_copy.get_sheet(sheet_idx)

    # compute regex column index
    if isinstance(regex_col_spec, (int,)) or (isinstance(regex_col_spec, float) and str(regex_col_spec).isdigit()):
        col_index = int(regex_col_spec)
    else:
        # string: either letter or header name
        rs = str(regex_col_spec).strip()
        if rs.isalpha():
            col_index = excel_col_to_index(rs)
        else:
            idx = find_header_index_xls(sheet, rs)
            if idx is None:
                raise ValueError(f"Regex column '{regex_col_spec}' not found in .xls header")
            col_index = idx

    pattern = changes.get("regex", "")
    try:
        regex = re.compile(pattern)
    except Exception as e:
        raise ValueError(f"Invalid regex pattern: {pattern}: {e}")

    # build header list for lookup of change columns
    headers = [str(h).strip() if h is not None else "" for h in sheet.row_values(0)]

    # iterate rows (skip header row 0)
    for r in range(1, sheet.nrows):
        cell_val = sheet.cell_value(r, col_index)
        if cell_val is None:
            continue
        if regex.search(str(cell_val)):
            # apply each change
            for col_spec, new_val in changes.get("changes", {}).items():
                # determine column index to write to
                if isinstance(col_spec, (int,)) or (isinstance(col_spec, float) and float(col_spec).is_integer()):
                    tgt_idx = int(col_spec)
                else:
                    cs = str(col_spec).strip()
                    if cs.isalpha():
                        tgt_idx = excel_col_to_index(cs)
                    else:
                        # treat as header name
                        if cs in headers:
                            tgt_idx = headers.index(cs)
                        else:
                            logging.warning(f"Change target column '{col_spec}' not found in .xls header; skipping.")
                            continue
                # write value (xlwt write)
                # If new_val is NaN from pandas, convert to empty
                if pd.isna(new_val):
                    write_val = ""
                else:
                    write_val = new_val
                sheet_writable.write(r, tgt_idx, write_val)

    return wb_copy

def process_xlsx(input_path, sheet_spec, regex_col_spec, changes):
    """
    Use openpyxl to edit .xlsx in-place (preserving styles).
    sheet_spec: int or sheet name
    regex_col_spec: header name, letter, or numeric index
    """
    wb = load_workbook(input_path)
    # resolve sheet
    if isinstance(sheet_spec, int):
        if sheet_spec < 0 or sheet_spec >= len(wb.worksheets):
            raise ValueError(f"Sheet index {sheet_spec} out of range in {input_path}")
        ws = wb.worksheets[sheet_spec]
    else:
        if sheet_spec not in wb.sheetnames:
            raise ValueError(f"Sheet name '{sheet_spec}' not found in {input_path}")
        ws = wb[sheet_spec]

    # header row values
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]

    # determine regex column index
    if isinstance(regex_col_spec, (int,)) or (isinstance(regex_col_spec, float) and float(regex_col_spec).is_integer()):
        col_index = int(regex_col_spec)
    else:
        rc = str(regex_col_spec).strip()
        if rc.isalpha():
            col_index = excel_col_to_index(rc)
        else:
            if rc in headers:
                col_index = headers.index(rc)
            else:
                raise ValueError(f"Regex column '{regex_col_spec}' not found in .xlsx header")

    pattern = changes.get("regex", "")
    try:
        regex = re.compile(pattern)
    except Exception as e:
        raise ValueError(f"Invalid regex pattern: {pattern}: {e}")

    # iterate rows starting from row 2
    for row_cells in ws.iter_rows(min_row=2):
        cell_obj = row_cells[col_index]
        cell_val = cell_obj.value
        if cell_val is None:
            continue
        if regex.search(str(cell_val)):
            for col_spec, new_val in changes.get("changes", {}).items():
                # resolve target col index
                if isinstance(col_spec, (int,)) or (isinstance(col_spec, float) and float(col_spec).is_integer()):
                    tgt_idx = int(col_spec)
                else:
                    cs = str(col_spec).strip()
                    if cs.isalpha():
                        tgt_idx = excel_col_to_index(cs)
                    else:
                        if cs in headers:
                            tgt_idx = headers.index(cs)
                        else:
                            logging.warning(f"Change target column '{col_spec}' not found in .xlsx header; skipping.")
                            continue
                # set value; cell objects in row_cells are ordered by columns starting at 0
                try:
                    target_cell = row_cells[tgt_idx]
                    target_cell.value = "" if pd.isna(new_val) else new_val
                except IndexError:
                    # If row_cells shorter than target idx (rare), write via ws.cell
                    ws.cell(row=cell_obj.row, column=(tgt_idx + 1), value = ("" if pd.isna(new_val) else new_val))

    return wb

def build_changes_from_rule_row(row, rules_columns):
    """
    Build changes dict {'regex': pattern, 'changes': {col_spec: value, ...}}
    The rules file is expected to have pairs like Change1_Column / Change1_Value, Change2_Column / Change2_Value, ...
    This scans available columns in the rules header to collect them.
    """
    changes = {"regex": None, "changes": {}}
    # regex is expected in 'Regex' column
    changes["regex"] = row.get("Regex", "")
    # find pairs by scanning column names
    # Accept patterns: Change{N}_Column and Change{N}_Value (case-insensitive)
    for col in rules_columns:
        c = str(col)
        if c.lower().startswith("change") and "_column" in c.lower():
            # extract suffix number
            prefix = c[:c.lower().find("_column")]
            value_col = prefix + "_Value"
            try:
                col_spec = row[col]
                if pd.isna(col_spec):
                    continue
                if value_col in rules_columns:
                    val = row[value_col]
                    if pd.isna(val):
                        continue
                    changes["changes"][col_spec] = val
            except Exception:
                continue
    return changes

def main():
    base = get_base_path()
    # rules filename — change here if you wish. We'll accept either rules.xls or rules.xlsx present in same folder.
    # Prefer rules.xls if exists, else rules.xlsx
    rules_xls = os.path.join(base, "rules.xls")
    rules_xlsx = os.path.join(base, "rules.xlsx")
    if os.path.exists(rules_xls):
        rules_file = rules_xls
    elif os.path.exists(rules_xlsx):
        rules_file = rules_xlsx
    else:
        logging.error("No rules.xls or rules.xlsx found in the script directory.")
        return

    logging.info(f"Reading rules from: {rules_file}")
    try:
        rules_df = read_rules(rules_file)
    except Exception as e:
        logging.error(f"Failed to read rules file: {e}")
        return

    # Ensure expected columns exist roughly (we'll be tolerant)
    # Process each grouped rule set by the 4 key columns if present, otherwise row-by-row
    required_keys = ["Input_File", "Input_Sheet", "Regex", "Regex_Column", "Output_File", "Output_Sheet"]
    missing = [k for k in required_keys if k not in rules_df.columns]
    if missing:
        logging.error(f"Rules file missing required columns: {missing}")
        return

    # iterate rows (we will group by the 4 keys just like you had before)
    grouped = rules_df.groupby(["Input_File", "Input_Sheet", "Output_File", "Output_Sheet"])
    for (input_file, input_sheet_raw, output_file, output_sheet_raw), group_df in grouped:
        input_file = str(input_file).strip()
        output_file = str(output_file).strip()
        input_path = os.path.join(base, input_file)
        output_path = os.path.join(base, output_file)

        if not os.path.exists(input_path):
            logging.error(f"Input file not found: {input_path}")
            continue

        input_sheet = normalize_sheet_spec(input_sheet_raw)
        output_sheet = normalize_sheet_spec(output_sheet_raw)

        # Build rules list for this group: each row can define one regex + multiple change pairs
        # We'll iterate group_df rows
        for _, rule_row in group_df.iterrows():
            # regex column spec (which column to test) can be header name or letter or index
            regex_col_spec = rule_row.get("Regex_Column", "")
            # build changes mapping from ChangeN_Column / ChangeN_Value columns in rules file
            changes = build_changes_from_rule_row(rule_row, rules_df.columns)
            if not changes["regex"]:
                logging.warning("Empty regex in rule; skipping.")
                continue

            # Determine ext (use input file extension to choose xls/xlsx path)
            ext = os.path.splitext(input_path)[1].lower()
            same_file = (os.path.normcase(input_path) == os.path.normcase(output_path))
            tmp_path = None
            try:
                if same_file:
                    # keep extension in the tmp filename to ensure correct engine
                    tmp_path = input_path.replace(ext, f".tmp{ext}")

                if ext == ".xls":
                    wb_copy = process_xls(input_path, input_sheet, regex_col_spec, changes)
                    save_to = tmp_path if same_file else output_path
                    wb_copy.save(save_to)
                    logging.info(f"Saved .xls to {save_to}")

                elif ext == ".xlsx":
                    wb = process_xlsx(input_path, input_sheet, regex_col_spec, changes)
                    save_to = tmp_path if same_file else output_path
                    wb.save(save_to)
                    logging.info(f"Saved .xlsx to {save_to}")

                else:
                    logging.error(f"Unsupported input extension: {ext}")
                    continue

                if same_file:
                    # replace original atomically
                    shutil.move(tmp_path, input_path)
                    logging.info(f"Replaced original file with updated: {input_path}")

            except Exception as e:
                logging.error(f"Failed processing {input_path} (rule regex={changes.get('regex')}): {e}")
                # continue with next rule

    logging.info("All rules processed.")

if __name__ == "__main__":
    main()
